from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import load_workbook

REVIEW_GROUP_KEY = "review_group_key"
ROW_NUMBER = "row_number"


def parse_group_assignments(values: list[str]) -> dict[int, str]:
    assignments: dict[int, str] = {}
    errors: list[str] = []
    for value in values:
        if "=" not in value:
            errors.append(f"expected row_number=review_group_key: {value}")
            continue
        row_text, group_key = value.split("=", 1)
        try:
            row_number = int(row_text.strip())
        except ValueError:
            errors.append(f"invalid row number: {row_text}")
            continue
        group_key = group_key.strip()
        if not group_key:
            errors.append(f"missing review_group_key for row {row_number}")
            continue
        assignments[row_number] = group_key
    if errors:
        raise ValueError("; ".join(errors))
    return assignments


def copy_with_review_group_keys(
    source: str | Path,
    output: str | Path,
    assignments: dict[int, str],
) -> None:
    source_path = Path(source)
    output_path = Path(output)
    workbook = load_workbook(source_path)
    try:
        worksheet = workbook.active
        headers = [str(cell.value or "").strip() for cell in worksheet[1]]
        if ROW_NUMBER not in headers:
            raise ValueError("review workbook missing required column: row_number")
        if REVIEW_GROUP_KEY not in headers:
            insert_at = headers.index("review_decision") + 1 if "review_decision" in headers else len(headers) + 1
            worksheet.insert_cols(insert_at)
            worksheet.cell(row=1, column=insert_at).value = REVIEW_GROUP_KEY
            headers.insert(insert_at - 1, REVIEW_GROUP_KEY)

        row_col = headers.index(ROW_NUMBER) + 1
        group_col = headers.index(REVIEW_GROUP_KEY) + 1
        rows_by_number = {
            int(worksheet.cell(row=row_index, column=row_col).value): row_index
            for row_index in range(2, worksheet.max_row + 1)
            if worksheet.cell(row=row_index, column=row_col).value is not None
        }
        unknown = sorted(row for row in assignments if row not in rows_by_number)
        if unknown:
            raise ValueError("unknown row_number: " + ", ".join(str(row) for row in unknown))

        for row_number, group_key in assignments.items():
            worksheet.cell(row=rows_by_number[row_number], column=group_col).value = group_key

        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output_path)
    finally:
        workbook.close()


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", required=True)
    parser.add_argument("--output", required=True)
    parser.add_argument("--set-group", action="append", default=[])
    args = parser.parse_args()
    assignments = parse_group_assignments(args.set_group)
    copy_with_review_group_keys(args.input, args.output, assignments)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
