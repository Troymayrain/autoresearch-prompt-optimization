from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Sequence

from openpyxl import load_workbook

REQUIRED_COLUMNS = (
    "group_key",
    "row_number",
    "review_group_key",
    "review_decision",
)
PROMPT_FIXABLE = "prompt_fixable"
EXTRA_CODE_OUTPUT = "extra_code_output"
REVIEW_GROUP_PRIORITY = (
    "wrong_code_selected_non_redeemable_number",
    "wrong_code_ocr_confusion",
    "no_card_false_negative",
    "missing_code",
    "extra_code_security_pin",
    "extra_code_barcode_receipt_number",
    "extra_code_output",
)


def build_review_feedback(
    review_workbook: str | Path,
    feedback_failures: dict[str, Any],
    dev_row_numbers: Sequence[int],
) -> dict[str, Any]:
    reviewed_rows = _read_review_rows(Path(review_workbook))
    dev_rows = {int(row) for row in dev_row_numbers}
    missing_dev_rows = sorted(row["row_number"] for row in reviewed_rows if row["row_number"] not in dev_rows)
    if missing_dev_rows:
        raise ValueError("review rows not found in current dev set: " + _rows_text(missing_dev_rows))

    row_feedback = _row_feedback(feedback_failures)
    active: dict[str, dict[str, Any]] = {}
    excluded_rows: list[dict[str, Any]] = []
    already_resolved_rows: list[dict[str, Any]] = []
    mismatched_rows: list[dict[str, Any]] = []
    missing_group_key_rows: list[int] = []

    for row in reviewed_rows:
        decision = row["review_decision"]
        original_key = row["group_key"]
        review_key = row["review_group_key"] or original_key
        if decision != PROMPT_FIXABLE:
            excluded_rows.append(
                {
                    "row_number": row["row_number"],
                    "group_key": original_key,
                    "review_decision": decision,
                }
            )
            continue
        if original_key == EXTRA_CODE_OUTPUT and not row["review_group_key"]:
            missing_group_key_rows.append(row["row_number"])
            continue

        current = row_feedback.get(row["row_number"])
        if current is None:
            already_resolved_rows.append(
                {
                    "row_number": row["row_number"],
                    "group_key": review_key,
                    "original_group_key": original_key,
                }
            )
            continue
        if not _compatible_group(original_key, review_key, current["group_key"]):
            mismatched_rows.append(
                {
                    "row_number": row["row_number"],
                    "group_key": review_key,
                    "original_group_key": original_key,
                    "current_group_key": current["group_key"],
                }
            )
            continue
        _add_active_row(active, review_key, row, current)

    if missing_group_key_rows:
        raise ValueError(
            "review_group_key required for prompt_fixable extra_code_output rows: "
            + _rows_text(missing_group_key_rows)
        )

    return {
        "task": feedback_failures.get("task"),
        "feedback_set": feedback_failures.get("feedback_set"),
        "review_workbook": str(review_workbook),
        "active_groups": _ordered_groups(active.values()),
        "background_groups": {
            "primary_groups": list(feedback_failures.get("primary_groups", [])),
            "secondary_groups": list(feedback_failures.get("secondary_groups", [])),
        },
        "excluded_rows": excluded_rows,
        "already_resolved_rows": already_resolved_rows,
        "mismatched_rows": mismatched_rows,
    }


def write_review_feedback(run_dir: str | Path, payload: dict[str, Any]) -> None:
    path = Path(run_dir)
    path.mkdir(parents=True, exist_ok=True)
    (path / "review-feedback.json").write_text(
        json.dumps(payload, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )


def _read_review_rows(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        rows = list(worksheet.iter_rows(values_only=True))
    finally:
        workbook.close()
    if not rows:
        raise ValueError("review workbook is empty")
    headers = [str(value or "").strip() for value in rows[0]]
    missing = [column for column in REQUIRED_COLUMNS if column not in headers]
    if missing:
        raise ValueError("review workbook missing required columns: " + ", ".join(missing))
    index = {column: headers.index(column) for column in REQUIRED_COLUMNS}
    reviewed = []
    for raw in rows[1:]:
        row_number = _row_number(raw[index["row_number"]])
        if row_number is None:
            continue
        reviewed.append(
            {
                "group_key": _text(raw[index["group_key"]]),
                "row_number": row_number,
                "review_group_key": _text(raw[index["review_group_key"]]),
                "review_decision": _text(raw[index["review_decision"]]),
            }
        )
    return reviewed


def _row_feedback(feedback_failures: dict[str, Any]) -> dict[int, dict[str, Any]]:
    by_row: dict[int, dict[str, Any]] = {}
    for group in list(feedback_failures.get("primary_groups", [])) + list(
        feedback_failures.get("secondary_groups", [])
    ):
        group_key = _text(group.get("key"))
        examples = {
            _row_number(example.get("row_number")): example
            for example in group.get("examples", [])
            if isinstance(example, dict)
        }
        for row_number in group.get("rows", []):
            row = _row_number(row_number)
            if row is None:
                continue
            by_row[row] = {
                "group_key": group_key,
                "failure_category": group.get("failure_category", ""),
                "reason": group.get("reason", group_key),
                "example": examples.get(row, {"row_number": row}),
            }
    return by_row


def _add_active_row(
    groups: dict[str, dict[str, Any]],
    group_key: str,
    review_row: dict[str, Any],
    current: dict[str, Any],
) -> None:
    group = groups.setdefault(
        group_key,
        {
            "key": group_key,
            "failure_category": current.get("failure_category", ""),
            "reason": current.get("reason", group_key),
            "count": 0,
            "rows": [],
            "examples": [],
        },
    )
    group["count"] += 1
    group["rows"].append(review_row["row_number"])
    example = dict(current.get("example") or {"row_number": review_row["row_number"]})
    example["review_group_key"] = group_key
    example["original_group_key"] = review_row["group_key"]
    if len(group["examples"]) < 5:
        group["examples"].append(example)


def _compatible_group(original_key: str, review_key: str, current_key: str) -> bool:
    return current_key in {original_key, review_key}


def _ordered_groups(groups: Sequence[dict[str, Any]]) -> list[dict[str, Any]]:
    by_key = {str(group.get("key", "")): group for group in groups}
    ordered = [by_key[key] for key in REVIEW_GROUP_PRIORITY if key in by_key]
    ordered.extend(group for key, group in by_key.items() if key not in REVIEW_GROUP_PRIORITY)
    return ordered


def _row_number(value: object) -> int | None:
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def _text(value: object) -> str:
    return "" if value is None else str(value).strip()


def _rows_text(rows: Sequence[int]) -> str:
    return ", ".join(str(row) for row in sorted(rows))
