from __future__ import annotations

import random
from dataclasses import dataclass
from pathlib import Path
from typing import Literal, Sequence

from openpyxl import load_workbook

TaskName = Literal["code", "type"]

CODE_REQUIRED_COLUMNS = ("card_image", "origin", "md5_card_number")
TYPE_REQUIRED_COLUMNS = ("card_image", "origin", "golden_type")
TYPE_VALUES = ("Physics", "E-codes")


@dataclass(frozen=True)
class Sample:
    row_number: int
    card_image: str
    origin: int
    expected_raw: str
    scoreable: bool


@dataclass(frozen=True)
class DatasetSplit:
    dev: list[Sample]
    full: list[Sample]


def _required_columns(task: TaskName) -> tuple[str, ...]:
    if task == "code":
        return CODE_REQUIRED_COLUMNS
    if task == "type":
        return TYPE_REQUIRED_COLUMNS
    raise ValueError(f"unsupported task: {task}")


def _header_map(values: Sequence[object], required_columns: Sequence[str]) -> dict[str, int]:
    found: dict[str, int] = {}
    for idx, value in enumerate(values):
        key = str(value or "").strip()
        if key:
            found[key] = idx
    missing = [name for name in required_columns if name not in found]
    if missing:
        raise ValueError(f"missing required columns: {', '.join(missing)}")
    return found


def _origin(value: object, row_number: int) -> int:
    try:
        return int(str(value).strip())
    except (TypeError, ValueError):
        raise ValueError(f"row {row_number} has invalid origin: {value!r}") from None


def _image_count(card_image: str) -> int:
    return len([part for part in card_image.split("||") if part.strip()])


def _expected_raw(
    row: Sequence[object],
    columns: dict[str, int],
    row_number: int,
    task: TaskName,
    image_count: int,
) -> str:
    column = "md5_card_number" if task == "code" else "golden_type"
    value = str(row[columns[column]] or "").strip()
    if task == "type" and value not in [type_value * image_count for type_value in TYPE_VALUES]:
        raise ValueError(
            f"row {row_number} has invalid golden_type: expected Physics or E-codes repeated {image_count} times"
        )
    return value


def load_dataset(path: str | Path, task: TaskName = "code") -> list[Sample]:
    workbook = load_workbook(Path(path), read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        rows = worksheet.iter_rows(values_only=True)
        headers = next(rows, ())

        columns = _header_map(headers, _required_columns(task))
        samples: list[Sample] = []
        for row_index, row in enumerate(rows, start=2):
            card_image = str(row[columns["card_image"]] or "").strip()
            if not card_image:
                raise ValueError(f"row {row_index} has empty card_image")
            expected_raw = _expected_raw(row, columns, row_index, task, _image_count(card_image))
            samples.append(
                Sample(
                    row_number=row_index,
                    card_image=card_image,
                    origin=_origin(row[columns["origin"]], row_index),
                    expected_raw=expected_raw,
                    scoreable=bool(expected_raw),
                )
            )
        return samples
    finally:
        workbook.close()


def split_samples(samples: Sequence[Sample], dev_size: int, seed: int = 20260507) -> DatasetSplit:
    full = list(samples)
    shuffled = list(full)
    random.Random(seed).shuffle(shuffled)
    return DatasetSplit(dev=shuffled[: min(dev_size, len(shuffled))], full=full)
