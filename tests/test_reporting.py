import json

from openpyxl import load_workbook

from optimizer.dataset import Sample
from optimizer.evaluation import EvaluationResult
from optimizer.reporting import write_feedback_failures, write_gate_artifact, write_run_artifacts


def _xlsx_headers(path):
    wb = load_workbook(path, read_only=True)
    try:
        return list(next(wb.active.iter_rows(values_only=True)))
    finally:
        wb.close()


def _xlsx_rows(path):
    wb = load_workbook(path, read_only=True)
    try:
        return list(wb.active.iter_rows(values_only=True))
    finally:
        wb.close()


def test_write_run_artifacts_creates_summary_failures_and_excel(tmp_path):
    sample = Sample(2, "a.png", 0, "ABC", True)
    result = EvaluationResult.from_ocr_response(sample, {"status": 200, "data": [], "imageStatus": ["ok"]})

    write_run_artifacts(
        run_dir=tmp_path,
        phase="dev",
        results=[result],
        prompt_before="old",
        prompt_after="new",
        optimizer_request={"a": 1},
        optimizer_response={"b": 2},
    )

    summary = json.loads((tmp_path / "summary.json").read_text())
    assert summary["task"] == "code"
    assert summary["phase"] == "dev"
    assert "business_accuracy" in summary
    assert (tmp_path / "failures.jsonl").read_text().strip()
    assert json.loads((tmp_path / "failures.jsonl").read_text())["task"] == "code"
    assert json.loads((tmp_path / "failure-clusters.json").read_text())["task"] == "code"
    assert (tmp_path / "results.xlsx").exists()
    assert not (tmp_path / "gate.json").exists()
    assert _xlsx_headers(tmp_path / "results.xlsx") == [
        "task",
        "row_number",
        "card_image",
        "origin",
        "expected",
        "actual",
        "business_correct",
        "business_total",
        "failure_category",
        "image_status",
    ]
    assert (tmp_path / "prompt.diff").read_text().startswith("--- prompt-before.js")


def test_write_feedback_failures_groups_primary_and_secondary_code_failures(tmp_path):
    results = [
        EvaluationResult.from_ocr_response(
            Sample(2, "number.png", 0, "W0B053BJVLF9QJL", True),
            {"status": 200, "data": [{"number": "6338730878581133"}], "imageStatus": ["ok"]},
        ),
        EvaluationResult.from_ocr_response(
            Sample(3, "ocr.png", 0, "F3Z9ZAZWEJXQYA8", True),
            {"status": 200, "data": [{"number": "F3Z9Z-AZWEJ-XQYAB"}], "imageStatus": ["ok"]},
        ),
        EvaluationResult.from_ocr_response(
            Sample(4, "no-card.png", 0, "A6TEXPIABVJ9L7Z", True),
            {"status": 200, "data": [], "imageStatus": ["no-card"]},
        ),
        EvaluationResult.from_ocr_response(
            Sample(5, "extra.png", 0, "ABC", True),
            {"status": 200, "data": [{"number": "ABC"}, {"number": "EXTRA"}], "imageStatus": ["ok"]},
        ),
    ]

    last_candidate_results = [
        EvaluationResult.from_ocr_response(
            Sample(3, "ocr.png", 0, "F3Z9ZAZWEJXQYA8", True),
            {"status": 200, "data": [{"number": "F3Z9Z-AZWEJ-XQYA8"}], "imageStatus": ["ok"]},
        )
    ]

    payload = write_feedback_failures(
        tmp_path,
        "dev",
        results,
        task="code",
        last_candidate_results=last_candidate_results,
    )

    written = json.loads((tmp_path / "feedback-failures.json").read_text())
    assert written == payload
    assert written["feedback_set"] == "dev"
    assert [group["key"] for group in written["primary_groups"]] == [
        "wrong_code_selected_non_redeemable_number",
        "wrong_code_ocr_confusion",
        "no_card_false_negative",
    ]
    assert written["primary_groups"][0]["rows"] == [2]
    assert written["primary_groups"][1]["rows"] == [3]
    assert written["primary_groups"][2]["rows"] == [4]
    assert [group["key"] for group in written["secondary_groups"]] == ["extra_code_output"]
    assert written["primary_groups"][0]["examples"][0]["origin"] == 0
    assert written["primary_groups"][0]["examples"][0]["card_image"] == "number.png"

    workbook_rows = _xlsx_rows(tmp_path / "feedback-review.xlsx")
    assert workbook_rows[0] == (
        "group_key",
        "row_number",
        "origin",
        "card_image",
        "expected",
        "accepted_actual",
        "last_candidate_actual",
        "failure_category",
        "review_decision",
        "review_notes",
    )
    assert workbook_rows[1:] == [
        (
            "wrong_code_selected_non_redeemable_number",
            2,
            0,
            "number.png",
            "W0B053BJVLF9QJL",
            "6338730878581133",
            None,
            "wrong_code",
            None,
            None,
        ),
        (
            "wrong_code_ocr_confusion",
            3,
            0,
            "ocr.png",
            "F3Z9ZAZWEJXQYA8",
            "F3Z9Z-AZWEJ-XQYAB",
            "F3Z9Z-AZWEJ-XQYA8",
            "wrong_code",
            None,
            None,
        ),
        (
            "no_card_false_negative",
            4,
            0,
            "no-card.png",
            "A6TEXPIABVJ9L7Z",
            None,
            None,
            "no_card",
            None,
            None,
        ),
        (
            "extra_code_output",
            5,
            0,
            "extra.png",
            "ABC",
            "ABC\nEXTRA",
            None,
            "extra_code",
            None,
            None,
        ),
    ]


def test_write_run_artifacts_redacts_optimizer_secrets(tmp_path):
    sample = Sample(2, "a.png", 0, "ABC", True)
    result = EvaluationResult.from_ocr_response(
        sample,
        {"status": 200, "data": [{"number": "ABC"}], "imageStatus": ["ok"]},
    )

    write_run_artifacts(
        run_dir=tmp_path,
        phase="dev",
        results=[result],
        prompt_before="old",
        prompt_after="new",
        optimizer_request={
            "normal": "keep me",
            "env": {"AI_GATEWAY_KEY": "secret-value", "DEC_SALT_TB": "salt-value"},
        },
        optimizer_response={"message": "ok"},
    )

    written = json.loads((tmp_path / "optimizer-request.json").read_text())
    assert written["normal"] == "keep me"
    assert written["env"]["AI_GATEWAY_KEY"] == "[REDACTED]"
    assert written["env"]["DEC_SALT_TB"] == "[REDACTED]"


def test_write_type_run_artifacts_uses_type_summary(tmp_path):
    mismatch = EvaluationResult.from_ocr_response(
        Sample(2, "a.png", 0, "Physics", True),
        {"status": 200, "data": [{"type": "E-codes", "number": "Physics"}], "imageStatus": ["ok"]},
        task="type",
    )
    missing = EvaluationResult.from_ocr_response(
        Sample(3, "b.png", 0, "Physics", True),
        {"status": 200, "data": [{"number": "Physics"}], "imageStatus": ["ok"]},
        task="type",
    )

    write_run_artifacts(
        run_dir=tmp_path,
        phase="full",
        results=[mismatch, missing],
        prompt_before="old",
        prompt_after="new",
        optimizer_request={},
        optimizer_response={},
        task="type",
    )

    summary = json.loads((tmp_path / "summary.json").read_text())
    assert summary["task"] == "type"
    assert summary["type_total"] == 1
    assert summary["type_correct"] == 0
    assert summary["type_accuracy"] == 0.0
    assert summary["evaluable_count"] == 1
    assert summary["not_evaluable_count"] == 1
    assert summary["failure_categories"] == {"not_evaluable": 1, "type_mismatch": 1}
    assert _xlsx_headers(tmp_path / "results.xlsx") == [
        "task",
        "row_number",
        "card_image",
        "origin",
        "expected",
        "actual",
        "type_correct",
        "type_total",
        "not_evaluable_reason",
        "failure_category",
        "image_status",
    ]


def test_write_gate_artifact_persists_decision_and_redacts_secrets(tmp_path):
    write_gate_artifact(
        tmp_path,
        {
            "task": "code",
            "phase": "regression",
            "decision": "discard",
            "checks": [{"name": "business_accuracy_not_decreased", "passed": False}],
            "reason": "business_accuracy decreased from 100.0 to 50.0",
            "metrics": {
                "accepted": {"business_accuracy": 100.0},
                "candidate": {"business_accuracy": 50.0, "api_key": "secret-value"},
            },
            "authorization": "Bearer secret",
        },
    )

    gate = json.loads((tmp_path / "gate.json").read_text())
    assert gate["task"] == "code"
    assert gate["phase"] == "regression"
    assert gate["decision"] == "discard"
    assert gate["checks"] == [{"name": "business_accuracy_not_decreased", "passed": False}]
    assert gate["reason"] == "business_accuracy decreased from 100.0 to 50.0"
    assert gate["metrics"]["accepted"]["business_accuracy"] == 100.0
    assert gate["metrics"]["candidate"]["api_key"] == "[REDACTED]"
    assert gate["authorization"] == "[REDACTED]"
