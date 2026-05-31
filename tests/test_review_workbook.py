from openpyxl import Workbook, load_workbook

from optimizer.review_workbook import copy_with_review_group_keys, parse_group_assignments


def _workbook(path):
    wb = Workbook()
    ws = wb.active
    ws.title = "feedback-review"
    ws.append(
        [
            "group_key",
            "row_number",
            "origin",
            "card_image",
            "expected",
            "accepted_actual",
            "last_candidate_actual",
            "failure_category",
            "review_decision",
            "review_notes",
        ]
    )
    ws.append(["extra_code_output", 113, 10, "a.png", "ABC", "ABC\nPIN", "", "extra_code", "prompt_fixable", ""])
    ws.append(["extra_code_output", 91, 10, "b.png", "DEF", "DEF\nBAR", "", "extra_code", "prompt_fixable", ""])
    wb.save(path)


def _rows(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        return list(wb.active.iter_rows(values_only=True))
    finally:
        wb.close()


def test_copy_with_review_group_keys_adds_column_and_sets_explicit_rows(tmp_path):
    source = tmp_path / "feedback-review.xlsx"
    output = tmp_path / "feedback-review.reviewed.xlsx"
    _workbook(source)

    copy_with_review_group_keys(
        source,
        output,
        {
            113: "extra_code_security_pin",
            91: "extra_code_barcode_receipt_number",
        },
    )

    rows = _rows(output)
    assert rows[0] == (
        "group_key",
        "row_number",
        "origin",
        "card_image",
        "expected",
        "accepted_actual",
        "last_candidate_actual",
        "failure_category",
        "review_group_key",
        "review_decision",
        "review_notes",
    )
    assert rows[1][8] == "extra_code_security_pin"
    assert rows[2][8] == "extra_code_barcode_receipt_number"
    assert _rows(source)[0][8] == "review_decision"


def test_copy_with_review_group_keys_rejects_unknown_rows(tmp_path):
    source = tmp_path / "feedback-review.xlsx"
    output = tmp_path / "feedback-review.reviewed.xlsx"
    _workbook(source)

    try:
        copy_with_review_group_keys(source, output, {999: "extra_code_security_pin"})
    except ValueError as exc:
        assert "unknown row_number: 999" in str(exc)
    else:
        raise AssertionError("expected unknown row_number to fail")


def test_parse_group_assignments_rejects_malformed_values():
    try:
        parse_group_assignments(["113", "abc=extra_code_security_pin"])
    except ValueError as exc:
        message = str(exc)
        assert "expected row_number=review_group_key" in message
        assert "invalid row number" in message
    else:
        raise AssertionError("expected malformed assignments to fail")
