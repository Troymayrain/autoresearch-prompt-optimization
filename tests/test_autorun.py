import argparse
import json
import re
import types

import pytest
from openpyxl import Workbook, load_workbook

import optimizer.autorun as autorun
from optimizer.autorun import _accuracy, should_stop, stop_reason
from optimizer.dataset import DatasetSplit, Sample
from optimizer.evaluation import EvaluationResult
from optimizer.llm import OptimizerProposal

CODE_RESULT_HEADERS = [
    "task",
    "row_number",
    "card_image",
    "origin",
    "expected",
    "actual",
    "business_correct",
    "business_total",
    "failure_category",
    "image_status",
]
TYPE_RESULT_HEADERS = [
    "task",
    "row_number",
    "card_image",
    "origin",
    "expected",
    "actual",
    "type_correct",
    "type_total",
    "not_evaluable_reason",
    "failure_category",
    "image_status",
]


def _xlsx_headers(path):
    wb = load_workbook(path, read_only=True)
    try:
        return list(next(wb.active.iter_rows(values_only=True)))
    finally:
        wb.close()


def _latest_session(tmp_path, task):
    task_root = tmp_path / "runs" / f"card-ocr-prompt-opt-{task}"
    latest = json.loads((task_root / "latest.json").read_text())
    return task_root / latest["session_dir"]


def _review_workbook(path, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "feedback-review"
    ws.append(
        [
            "group_key",
            "row_number",
            "origin",
            "card_image",
            "expected",
            "accepted_actual",
            "last_candidate_actual",
            "failure_category",
            "review_group_key",
            "review_decision",
            "review_notes",
        ]
    )
    for row in rows:
        ws.append(row)
    wb.save(path)


def test_stop_when_target_reached():
    assert should_stop(iteration=3, full_accuracy=99.0, target=99.0, plateau_count=0, plateau_window=3, max_iterations=15)


def test_stop_when_plateau_window_reached():
    assert should_stop(iteration=4, full_accuracy=90.0, target=99.0, plateau_count=3, plateau_window=3, max_iterations=15)


def test_stop_when_max_iterations_reached():
    assert should_stop(iteration=15, full_accuracy=90.0, target=99.0, plateau_count=0, plateau_window=3, max_iterations=15)


def test_stop_reason_priority():
    assert (
        stop_reason(
            iteration=15,
            full_accuracy=99.0,
            target=99.0,
            plateau_count=3,
            plateau_window=3,
            max_iterations=15,
            no_business_learning_count=3,
            no_business_learning_window=3,
        )
        == "target_reached"
    )
    assert (
        stop_reason(
            iteration=15,
            full_accuracy=90.0,
            target=99.0,
            plateau_count=3,
            plateau_window=3,
            max_iterations=15,
            no_business_learning_count=3,
            no_business_learning_window=3,
        )
        == "no_business_learning"
    )
    assert (
        stop_reason(
            iteration=15,
            full_accuracy=90.0,
            target=99.0,
            plateau_count=3,
            plateau_window=3,
            max_iterations=15,
            no_business_learning_count=0,
            no_business_learning_window=3,
        )
        == "plateau"
    )


def test_continue_before_limits():
    assert not should_stop(iteration=2, full_accuracy=90.0, target=99.0, plateau_count=1, plateau_window=3, max_iterations=15)


def test_create_session_dir_adds_suffix_when_timestamp_collides(tmp_path, monkeypatch):
    monkeypatch.setattr(autorun, "_session_timestamp", lambda: "2026-05-30_12-00-00")
    task_root = tmp_path / "runs" / "card-ocr-prompt-opt-code"
    (task_root / "2026-05-30_12-00-00").mkdir(parents=True)

    session = autorun._create_session_dir(tmp_path / "runs", "code")

    latest = json.loads((task_root / "latest.json").read_text())
    assert session == task_root / "2026-05-30_12-00-00-02"
    assert latest["session_dir"] == "2026-05-30_12-00-00-02"


def test_accuracy_uses_business_score_only():
    result = EvaluationResult.from_ocr_response(
        Sample(2, "a.png", 0, "A-B-C", True),
        {"status": 200, "data": [{"number": "ABC"}], "imageStatus": ["ok"]},
    )

    assert _accuracy([result], task="code") == 100.0


def test_accuracy_uses_type_score_for_type_task():
    result = EvaluationResult.from_ocr_response(
        Sample(2, "a.png", 0, "Physics", True),
        {"status": 200, "data": [{"type": "Physics", "number": "wrong"}], "imageStatus": ["ok"]},
        task="type",
    )

    assert _accuracy([result], task="type") == 100.0


@pytest.mark.asyncio
async def test_autorun_creates_isolated_session_and_latest_pointer(tmp_path, monkeypatch, capsys):
    prompt = tmp_path / "prompts" / "ocr.js"
    prompt.parent.mkdir()
    prompt.write_text("valid", encoding="utf-8")
    task_root = tmp_path / "runs" / "card-ocr-prompt-opt-code"
    legacy_run = task_root / "run-999"
    legacy_run.mkdir(parents=True)
    sample = Sample(2, "a.png", 0, "A", True)
    cfg = types.SimpleNamespace(
        dev_sample_size=1,
        ocr_concurrency=1,
        runs_dir=tmp_path / "runs",
        prompt_path=prompt,
        node_binary="node",
        ocr_runner_path="runner.js",
        max_iterations=0,
        target_business_accuracy=101.0,
        plateau_window=3,
        optimizer_provider="test",
        optimizer_model="test",
    )

    monkeypatch.setattr(autorun.OptimizerConfig, "from_env", classmethod(lambda cls: cfg))
    monkeypatch.setattr(autorun, "load_dataset", lambda path, task: [sample])
    monkeypatch.setattr(autorun, "split_samples", lambda samples, size: DatasetSplit(dev=[sample], full=[sample]))
    monkeypatch.setattr(
        autorun,
        "validate_prompt_file",
        lambda path, node_binary="node", task=None, baseline_source=None: None,
    )

    async def fake_run_once(samples, runner, concurrency, task):
        return [
            EvaluationResult.from_ocr_response(
                item,
                {"status": 200, "data": [{"number": "A"}], "imageStatus": ["ok"]},
            )
            for item in samples
        ]

    monkeypatch.setattr(autorun, "run_once", fake_run_once)

    await autorun.main_async(argparse.Namespace(dataset="dataset.xlsx", task="code"))

    latest = json.loads((task_root / "latest.json").read_text())
    session_dir = latest["session_dir"]
    session = task_root / session_dir
    assert re.match(r"\d{4}-\d{2}-\d{2}_\d{2}-\d{2}-\d{2}(?:-\d{2})?$", session_dir)
    assert latest == {
        "task": "code",
        "session_dir": session_dir,
        "path": f"runs/card-ocr-prompt-opt-code/{session_dir}",
    }
    assert (session / "run-000-baseline" / "summary.json").exists()
    stop = json.loads((session / "stop.json").read_text())
    assert stop["reason"] == "max_iterations"
    assert stop["iteration"] == 0
    assert stop["last_run_dir"] == "run-000-baseline"
    assert stop["last_phase"] == "full"
    assert f"run_session=runs/card-ocr-prompt-opt-code/{session_dir}" in capsys.readouterr().out
    assert not (task_root / "run-000-baseline" / "summary.json").exists()
    assert legacy_run.exists()


@pytest.mark.asyncio
async def test_autorun_stops_after_no_business_learning_window(tmp_path, monkeypatch, capsys):
    prompt = tmp_path / "prompts" / "ocr.js"
    prompt.parent.mkdir()
    prompt.write_text("accepted", encoding="utf-8")
    sample = Sample(2, "a.png", 0, "A", True)
    cfg = types.SimpleNamespace(
        dev_sample_size=1,
        ocr_concurrency=1,
        runs_dir=tmp_path / "runs",
        prompt_path=prompt,
        node_binary="node",
        ocr_runner_path="runner.js",
        max_iterations=5,
        target_business_accuracy=101.0,
        plateau_window=99,
        no_business_learning_window=2,
        optimizer_provider="test",
        optimizer_model="test",
    )
    proposals = iter(
        [
            OptimizerProposal("h1", "e1", "r1", ["row 2"], "bad-one"),
            OptimizerProposal("h2", "e2", "r2", ["row 2"], "bad-two"),
        ]
    )
    seen_requests = []

    monkeypatch.setattr(autorun.OptimizerConfig, "from_env", classmethod(lambda cls: cfg))
    monkeypatch.setattr(autorun, "load_dataset", lambda path, task: [sample])
    monkeypatch.setattr(autorun, "split_samples", lambda samples, size: DatasetSplit(dev=[sample], full=[sample]))
    monkeypatch.setattr(
        autorun,
        "validate_prompt_file",
        lambda path, node_binary="node", task=None, baseline_source=None: None,
    )
    monkeypatch.setattr(autorun, "call_optimizer_llm", lambda provider, model, system, user: seen_requests.append(json.loads(user)) or next(proposals))
    monkeypatch.setattr(autorun, "commit_prompt", lambda path, message: pytest.fail("should not commit"))

    async def fake_run_once(samples, runner, concurrency, task):
        actual = "A" if prompt.read_text(encoding="utf-8") == "accepted" else "B"
        return [
            EvaluationResult.from_ocr_response(
                item,
                {"status": 200, "data": [{"number": actual}], "imageStatus": ["ok"]},
            )
            for item in samples
        ]

    monkeypatch.setattr(autorun, "run_once", fake_run_once)

    await autorun.main_async(argparse.Namespace(dataset="dataset.xlsx", task="code"))

    session = _latest_session(tmp_path, "code")
    stop = json.loads((session / "stop.json").read_text())
    assert len(seen_requests) == 2
    assert prompt.read_text(encoding="utf-8") == "accepted"
    assert stop["reason"] == "no_business_learning"
    assert stop["iteration"] == 2
    assert stop["last_run_dir"] == "run-002"
    assert stop["last_phase"] == "dev"
    assert stop["no_business_learning_count"] == 2
    assert "stop_reason=no_business_learning" in capsys.readouterr().out


@pytest.mark.asyncio
async def test_no_business_learning_stops_when_only_focused_group_regresses(tmp_path, monkeypatch):
    prompt = tmp_path / "prompts" / "ocr.js"
    prompt.parent.mkdir()
    prompt.write_text("accepted", encoding="utf-8")
    samples = [
        Sample(2, "a.png", 0, "A", True),
        Sample(3, "b.png", 0, "B", True),
    ]
    cfg = types.SimpleNamespace(
        dev_sample_size=2,
        ocr_concurrency=1,
        runs_dir=tmp_path / "runs",
        prompt_path=prompt,
        node_binary="node",
        ocr_runner_path="runner.js",
        max_iterations=6,
        target_business_accuracy=101.0,
        plateau_window=99,
        no_business_learning_window=2,
        optimizer_provider="test",
        optimizer_model="test",
    )
    proposals = iter(
        [
            OptimizerProposal("h1", "e1", "r1", ["row 2"], "bad-one"),
            OptimizerProposal("h2", "e2", "r2", ["row 3"], "mixed"),
            OptimizerProposal("h3", "e3", "r3", ["row 2"], "bad-three"),
            OptimizerProposal("h4", "e4", "r4", ["row 2"], "bad-four"),
        ]
    )
    seen_requests = []

    monkeypatch.setattr(autorun.OptimizerConfig, "from_env", classmethod(lambda cls: cfg))
    monkeypatch.setattr(autorun, "load_dataset", lambda path, task: samples)
    monkeypatch.setattr(autorun, "split_samples", lambda items, size: DatasetSplit(dev=items, full=items))
    monkeypatch.setattr(
        autorun,
        "validate_prompt_file",
        lambda path, node_binary="node", task=None, baseline_source=None: None,
    )
    monkeypatch.setattr(autorun, "call_optimizer_llm", lambda provider, model, system, user: seen_requests.append(json.loads(user)) or next(proposals))
    monkeypatch.setattr(autorun, "commit_prompt", lambda path, message: pytest.fail("should not commit"))

    async def fake_run_once(items, runner, concurrency, task):
        prompt_text = prompt.read_text(encoding="utf-8")
        actual_by_row = {
            "accepted": {2: "A", 3: ""},
            "mixed": {2: "MISS", 3: "B"},
        }.get(prompt_text, {2: "MISS", 3: "MISS"})
        return [
            EvaluationResult.from_ocr_response(
                item,
                {"status": 200, "data": [{"number": actual_by_row[item.row_number]}], "imageStatus": ["ok"]},
            )
            for item in items
        ]

    monkeypatch.setattr(autorun, "run_once", fake_run_once)

    await autorun.main_async(argparse.Namespace(dataset="dataset.xlsx", task="code"))

    stop = json.loads((_latest_session(tmp_path, "code") / "stop.json").read_text())
    assert len(seen_requests) == 1
    assert seen_requests[0]["focused_feedback_group"]["key"] == "missing_code"
    assert prompt.read_text(encoding="utf-8") == "accepted"
    assert stop["reason"] == "no_business_learning"
    assert stop["iteration"] == 1
    assert stop["no_business_learning_count"] == 2
    assert stop["last_phase"] == "focused_feedback_exhausted"


@pytest.mark.asyncio
async def test_autorun_uses_last_accepted_summary_after_rejected_candidate(tmp_path, monkeypatch):
    prompt = tmp_path / "prompts" / "ocr.js"
    prompt.parent.mkdir()
    prompt.write_text("good", encoding="utf-8")
    sample = Sample(2, "a.png", 0, "A", True)
    cfg = types.SimpleNamespace(
        dev_sample_size=1,
        ocr_concurrency=1,
        runs_dir=tmp_path / "runs",
        prompt_path=prompt,
        node_binary="node",
        ocr_runner_path="runner.js",
        max_iterations=2,
        target_business_accuracy=101.0,
        plateau_window=3,
        optimizer_provider="test",
        optimizer_model="test",
    )
    seen_accuracies = []
    seen_payloads = []
    proposals = iter(
        [
            OptimizerProposal("h", "e", "r", ["row 2"], "bad"),
            OptimizerProposal("h", "e", "r", ["row 2"], "good"),
        ]
    )

    monkeypatch.setattr(autorun.OptimizerConfig, "from_env", classmethod(lambda cls: cfg))
    loaded_tasks = []

    def fake_load_dataset(path, task):
        loaded_tasks.append(task)
        return [sample]

    monkeypatch.setattr(autorun, "load_dataset", fake_load_dataset)
    monkeypatch.setattr(autorun, "split_samples", lambda samples, size: DatasetSplit(dev=[sample], full=[sample]))
    monkeypatch.setattr(
        autorun,
        "validate_prompt_file",
        lambda path, node_binary="node", task=None, baseline_source=None: None,
    )
    monkeypatch.setattr(autorun, "commit_prompt", lambda path, message: pytest.fail("should not commit"))

    async def fake_run_once(samples, runner, concurrency, task):
        assert task == "code"
        actual = "B" if prompt.read_text(encoding="utf-8") == "bad" else "A"
        return [
            EvaluationResult.from_ocr_response(
                item,
                {"status": 200, "data": [{"number": actual}], "imageStatus": ["ok"]},
            )
            for item in samples
        ]

    def fake_call(provider, model, system, user):
        payload = json.loads(user)
        seen_payloads.append(payload)
        seen_accuracies.append(payload["summary"]["business_accuracy"])
        return next(proposals)

    monkeypatch.setattr(autorun, "run_once", fake_run_once)
    monkeypatch.setattr(autorun, "call_optimizer_llm", fake_call)

    await autorun.main_async(argparse.Namespace(dataset="dataset.xlsx", task="code"))

    assert loaded_tasks == ["code"]
    assert seen_accuracies == [100.0, 100.0]
    assert "candidate_evaluation_delta_summary" not in seen_payloads[0]
    assert seen_payloads[1]["candidate_evaluation_delta_summary"][0]["regressed_business_rows"] == [
        {
            "row_number": 2,
            "accepted_failure_category": "",
            "candidate_failure_category": "wrong_code",
            "business_delta": -1,
            "strict_delta": -1,
            "accepted_actual": ["A"],
            "candidate_actual": ["B"],
        }
    ]
    assert prompt.read_text(encoding="utf-8") == "good"
    delta = json.loads((_latest_session(tmp_path, "code") / "run-001/dev-delta.json").read_text())
    assert delta["regressed_business_rows"][0]["row_number"] == 2
    assert delta["target_failures_effect"]["row_targets"] == [
        {"target": "row 2", "row_number": 2, "outcome": "regressed"}
    ]


@pytest.mark.asyncio
async def test_autorun_uses_dev_feedback_failures_in_optimizer_request(tmp_path, monkeypatch):
    prompt = tmp_path / "prompts" / "ocr.js"
    prompt.parent.mkdir()
    prompt.write_text("accepted", encoding="utf-8")
    dev_sample = Sample(2, "dev.png", 0, "DEV-CODE", True)
    full_only_sample = Sample(99, "full.png", 0, "FULL-CODE", True)
    samples = [dev_sample, full_only_sample]
    cfg = types.SimpleNamespace(
        dev_sample_size=1,
        ocr_concurrency=1,
        runs_dir=tmp_path / "runs",
        prompt_path=prompt,
        node_binary="node",
        ocr_runner_path="runner.js",
        max_iterations=1,
        target_business_accuracy=101.0,
        plateau_window=3,
        optimizer_provider="test",
        optimizer_model="test",
    )
    seen_payloads = []

    monkeypatch.setattr(autorun.OptimizerConfig, "from_env", classmethod(lambda cls: cfg))
    monkeypatch.setattr(autorun, "load_dataset", lambda path, task: samples)
    monkeypatch.setattr(
        autorun,
        "split_samples",
        lambda items, size: DatasetSplit(dev=[dev_sample], full=items),
    )
    monkeypatch.setattr(
        autorun,
        "validate_prompt_file",
        lambda path, node_binary="node", task=None, baseline_source=None: None,
    )
    monkeypatch.setattr(autorun, "commit_prompt", lambda path, message: pytest.fail("should not commit"))

    async def fake_run_once(items, runner, concurrency, task):
        prompt_text = prompt.read_text(encoding="utf-8")
        actual_by_row = (
            {2: "6338730878581133", 99: "4251976762454"}
            if prompt_text == "accepted"
            else {2: "still-wrong", 99: "unused"}
        )
        return [
            EvaluationResult.from_ocr_response(
                item,
                {"status": 200, "data": [{"number": actual_by_row[item.row_number]}], "imageStatus": ["ok"]},
            )
            for item in items
        ]

    def fake_call(provider, model, system, user):
        seen_payloads.append(json.loads(user))
        return OptimizerProposal(
            "h",
            "e",
            "r",
            ["wrong_code_selected_non_redeemable_number"],
            "candidate",
        )

    monkeypatch.setattr(autorun, "run_once", fake_run_once)
    monkeypatch.setattr(autorun, "call_optimizer_llm", fake_call)

    await autorun.main_async(argparse.Namespace(dataset="dataset.xlsx", task="code"))

    payload = seen_payloads[0]
    assert payload["optimizer_background_evidence"]["summary"]["samples"] == 2
    assert "representative_failures" not in payload
    assert payload["optimizer_feedback_set"] == {"task": "code", "feedback_set": "dev"}
    assert payload["focused_feedback_group"]["key"] == "wrong_code_selected_non_redeemable_number"
    assert payload["focused_feedback_group"]["rows"] == [2]
    assert payload["optimizer_background_evidence"]["inactive_primary_groups"] == []
    assert "99" not in json.dumps(payload["focused_feedback_group"])
    feedback_path = _latest_session(tmp_path, "code") / "run-000-baseline" / "feedback-failures.json"
    assert json.loads(feedback_path.read_text())["primary_groups"][0]["rows"] == [2]
    delta = json.loads((_latest_session(tmp_path, "code") / "run-001/dev-delta.json").read_text())
    assert delta["target_failures_effect"]["row_targets"] == [
        {
            "target": "row 2",
            "row_number": 2,
            "outcome": "unchanged",
        }
    ]
    assert delta["target_failures_effect"]["category_targets"] == []


@pytest.mark.asyncio
async def test_autorun_review_workbook_promotes_reviewed_groups_to_focused_feedback(tmp_path, monkeypatch):
    prompt = tmp_path / "prompts" / "ocr.js"
    prompt.parent.mkdir()
    prompt.write_text("accepted", encoding="utf-8")
    workbook = tmp_path / "feedback-review.xlsx"
    _review_workbook(
        workbook,
        [
            [
                "extra_code_output",
                113,
                10,
                "pin.png",
                "ABC",
                "ABC\nPIN",
                "",
                "extra_code",
                "extra_code_security_pin",
                "prompt_fixable",
                "pin",
            ]
        ],
    )
    sample = Sample(113, "pin.png", 10, "ABC", True)
    cfg = types.SimpleNamespace(
        dev_sample_size=1,
        ocr_concurrency=1,
        runs_dir=tmp_path / "runs",
        prompt_path=prompt,
        node_binary="node",
        ocr_runner_path="runner.js",
        max_iterations=1,
        target_business_accuracy=101.0,
        plateau_window=3,
        optimizer_provider="test",
        optimizer_model="test",
    )
    seen_payloads = []

    monkeypatch.setattr(autorun.OptimizerConfig, "from_env", classmethod(lambda cls: cfg))
    monkeypatch.setattr(autorun, "load_dataset", lambda path, task: [sample])
    monkeypatch.setattr(autorun, "split_samples", lambda items, size: DatasetSplit(dev=items, full=items))
    monkeypatch.setattr(
        autorun,
        "validate_prompt_file",
        lambda path, node_binary="node", task=None, baseline_source=None: None,
    )
    monkeypatch.setattr(autorun, "commit_prompt", lambda path, message: pytest.fail("should not commit"))

    async def fake_run_once(items, runner, concurrency, task):
        return [
            EvaluationResult.from_ocr_response(
                item,
                {"status": 200, "data": [{"number": "ABC"}, {"number": "PIN"}], "imageStatus": ["ok"]},
            )
            for item in items
        ]

    def fake_call(provider, model, system, user):
        payload = json.loads(user)
        seen_payloads.append(payload)
        return OptimizerProposal("h", "e", "r", [payload["focused_feedback_group"]["key"]], "candidate")

    monkeypatch.setattr(autorun, "run_once", fake_run_once)
    monkeypatch.setattr(autorun, "call_optimizer_llm", fake_call)

    await autorun.main_async(
        argparse.Namespace(dataset="dataset.xlsx", task="code", review_workbook=str(workbook))
    )

    payload = seen_payloads[0]
    assert payload["focused_feedback_group"] == {"key": "extra_code_security_pin", "rows": [113]}
    assert payload["optimizer_background_evidence"]["inactive_primary_groups"] == []
    session = _latest_session(tmp_path, "code")
    review = json.loads((session / "run-000-baseline/review-feedback.json").read_text())
    assert review["active_groups"][0]["key"] == "extra_code_security_pin"
    delta = json.loads((session / "run-001/dev-delta.json").read_text())
    assert delta["reviewed_target_effect"]["summary"] == {
        "resolved": 0,
        "unchanged": 1,
        "regressed": 0,
        "ignored": 0,
    }


@pytest.mark.asyncio
async def test_autorun_rejects_review_workbook_for_type_before_loading_dataset(tmp_path, monkeypatch):
    monkeypatch.setattr(autorun, "load_dataset", lambda path, task: pytest.fail("should fail before loading dataset"))

    with pytest.raises(ValueError, match="--review-workbook is only supported for task code"):
        await autorun.main_async(
            argparse.Namespace(dataset="dataset.xlsx", task="type", review_workbook=str(tmp_path / "review.xlsx"))
        )


@pytest.mark.asyncio
async def test_autorun_records_failed_memory_and_stops_after_primary_groups_exhausted(tmp_path, monkeypatch):
    prompt = tmp_path / "prompts" / "ocr.js"
    prompt.parent.mkdir()
    prompt.write_text("accepted", encoding="utf-8")
    samples = [
        Sample(2, "number.png", 0, "W0B053BJVLF9QJL", True),
        Sample(3, "no-card.png", 0, "A6TEXPIABVJ9L7Z", True),
    ]
    cfg = types.SimpleNamespace(
        dev_sample_size=2,
        ocr_concurrency=1,
        runs_dir=tmp_path / "runs",
        prompt_path=prompt,
        node_binary="node",
        ocr_runner_path="runner.js",
        max_iterations=5,
        target_business_accuracy=101.0,
        plateau_window=5,
        no_business_learning_window=3,
        optimizer_provider="test",
        optimizer_model="test",
    )
    seen_payloads = []

    monkeypatch.setattr(autorun.OptimizerConfig, "from_env", classmethod(lambda cls: cfg))
    monkeypatch.setattr(autorun, "load_dataset", lambda path, task: samples)
    monkeypatch.setattr(
        autorun,
        "split_samples",
        lambda items, size: DatasetSplit(dev=items, full=items),
    )
    monkeypatch.setattr(
        autorun,
        "validate_prompt_file",
        lambda path, node_binary="node", task=None, baseline_source=None: None,
    )
    monkeypatch.setattr(autorun, "commit_prompt", lambda path, message: pytest.fail("should not commit"))

    async def fake_run_once(items, runner, concurrency, task):
        return [
            EvaluationResult.from_ocr_response(
                item,
                (
                    {"status": 200, "data": [{"number": "6338730878581133"}], "imageStatus": ["ok"]}
                    if item.row_number == 2
                    else {"status": 200, "data": [], "imageStatus": ["no-card"]}
                ),
            )
            for item in items
        ]

    def fake_call(provider, model, system, user):
        payload = json.loads(user)
        seen_payloads.append(payload)
        return OptimizerProposal(
            "same ineffective strategy",
            "no target improvement",
            "low",
            [payload["focused_feedback_group"]["key"]],
            f"candidate-{len(seen_payloads)}",
        )

    monkeypatch.setattr(autorun, "run_once", fake_run_once)
    monkeypatch.setattr(autorun, "call_optimizer_llm", fake_call)

    await autorun.main_async(argparse.Namespace(dataset="dataset.xlsx", task="code"))

    assert [payload["focused_feedback_group"]["key"] for payload in seen_payloads] == [
        "wrong_code_selected_non_redeemable_number",
        "no_card_false_negative",
    ]
    assert seen_payloads[1]["failed_strategy_memory"][0]["focused_group"] == (
        "wrong_code_selected_non_redeemable_number"
    )
    session = _latest_session(tmp_path, "code")
    stop = json.loads((session / "stop.json").read_text())
    assert stop["reason"] == "no_business_learning"
    assert stop["no_business_learning_count"] == 3
    memory = json.loads((session / "run-001/failed-strategy-memory.json").read_text())
    assert memory["entries"][0]["outcome"] == "unchanged"
    assert not (session / "run-003").exists()


@pytest.mark.asyncio
async def test_no_business_learning_waits_for_all_focused_primary_groups(tmp_path, monkeypatch):
    prompt = tmp_path / "prompts" / "ocr.js"
    prompt.parent.mkdir()
    prompt.write_text("accepted", encoding="utf-8")
    samples = [
        Sample(2, "number.png", 0, "W0B053BJVLF9QJL", True),
        Sample(3, "ocr.png", 0, "F3Z9ZAZWEJXQYA8", True),
        Sample(4, "no-card.png", 0, "A6TEXPIABVJ9L7Z", True),
        Sample(5, "missing.png", 0, "MISSING", True),
    ]
    cfg = types.SimpleNamespace(
        dev_sample_size=4,
        ocr_concurrency=1,
        runs_dir=tmp_path / "runs",
        prompt_path=prompt,
        node_binary="node",
        ocr_runner_path="runner.js",
        max_iterations=6,
        target_business_accuracy=101.0,
        plateau_window=99,
        optimizer_provider="test",
        optimizer_model="test",
    )
    seen_groups = []

    monkeypatch.setattr(autorun.OptimizerConfig, "from_env", classmethod(lambda cls: cfg))
    monkeypatch.setattr(autorun, "load_dataset", lambda path, task: samples)
    monkeypatch.setattr(autorun, "split_samples", lambda items, size: DatasetSplit(dev=items, full=items))
    monkeypatch.setattr(
        autorun,
        "validate_prompt_file",
        lambda path, node_binary="node", task=None, baseline_source=None: None,
    )
    monkeypatch.setattr(autorun, "commit_prompt", lambda path, message: pytest.fail("should not commit"))

    async def fake_run_once(items, runner, concurrency, task):
        actual_by_row = {
            2: "6338730878581133",
            3: "F3Z9Z-AZWEJ-XQYAB",
            4: "",
            5: "",
        }
        return [
            EvaluationResult.from_ocr_response(
                item,
                {
                    "status": 200,
                    "data": [{"number": actual_by_row[item.row_number]}]
                    if actual_by_row[item.row_number]
                    else [],
                    "imageStatus": ["no-card"] if item.row_number == 4 else ["ok"],
                },
            )
            for item in items
        ]

    def fake_call(provider, model, system, user):
        payload = json.loads(user)
        group_key = payload["focused_feedback_group"]["key"]
        seen_groups.append(group_key)
        return OptimizerProposal("h", "e", "r", [group_key], f"candidate-{len(seen_groups)}")

    monkeypatch.setattr(autorun, "run_once", fake_run_once)
    monkeypatch.setattr(autorun, "call_optimizer_llm", fake_call)

    await autorun.main_async(argparse.Namespace(dataset="dataset.xlsx", task="code"))

    assert seen_groups == [
        "wrong_code_selected_non_redeemable_number",
        "wrong_code_ocr_confusion",
        "no_card_false_negative",
        "missing_code",
    ]
    stop = json.loads((_latest_session(tmp_path, "code") / "stop.json").read_text())
    assert stop["reason"] == "no_business_learning"
    assert stop["iteration"] == 4
    assert stop["last_phase"] == "focused_feedback_exhausted"


@pytest.mark.asyncio
@pytest.mark.parametrize(
    ("max_iterations", "plateau_window", "expected_reason", "expected_count"),
    [
        (4, 2, "plateau", 2),
        (3, 99, "max_iterations", 3),
    ],
)
async def test_focused_runs_preserve_plateau_and_max_iteration_stop_reason(
    tmp_path,
    monkeypatch,
    max_iterations,
    plateau_window,
    expected_reason,
    expected_count,
):
    prompt = tmp_path / "prompts" / "ocr.js"
    prompt.parent.mkdir()
    prompt.write_text("accepted", encoding="utf-8")
    samples = [
        Sample(2, "number.png", 0, "W0B053BJVLF9QJL", True),
        Sample(3, "ocr.png", 0, "F3Z9ZAZWEJXQYA8", True),
        Sample(4, "no-card.png", 0, "A6TEXPIABVJ9L7Z", True),
        Sample(5, "missing.png", 0, "MISSING", True),
    ]
    cfg = types.SimpleNamespace(
        dev_sample_size=4,
        ocr_concurrency=1,
        runs_dir=tmp_path / "runs",
        prompt_path=prompt,
        node_binary="node",
        ocr_runner_path="runner.js",
        max_iterations=max_iterations,
        target_business_accuracy=101.0,
        plateau_window=plateau_window,
        no_business_learning_window=1,
        optimizer_provider="test",
        optimizer_model="test",
    )
    seen_groups = []

    monkeypatch.setattr(autorun.OptimizerConfig, "from_env", classmethod(lambda cls: cfg))
    monkeypatch.setattr(autorun, "load_dataset", lambda path, task: samples)
    monkeypatch.setattr(autorun, "split_samples", lambda items, size: DatasetSplit(dev=items, full=items))
    monkeypatch.setattr(
        autorun,
        "validate_prompt_file",
        lambda path, node_binary="node", task=None, baseline_source=None: None,
    )
    monkeypatch.setattr(autorun, "commit_prompt", lambda path, message: pytest.fail("should not commit"))

    async def fake_run_once(items, runner, concurrency, task):
        actual_by_row = {
            2: "6338730878581133",
            3: "F3Z9Z-AZWEJ-XQYAB",
            4: "",
            5: "",
        }
        return [
            EvaluationResult.from_ocr_response(
                item,
                {
                    "status": 200,
                    "data": [{"number": actual_by_row[item.row_number]}]
                    if actual_by_row[item.row_number]
                    else [],
                    "imageStatus": ["no-card"] if item.row_number == 4 else ["ok"],
                },
            )
            for item in items
        ]

    def fake_call(provider, model, system, user):
        payload = json.loads(user)
        group_key = payload["focused_feedback_group"]["key"]
        seen_groups.append(group_key)
        return OptimizerProposal("h", "e", "r", [group_key], f"candidate-{len(seen_groups)}")

    monkeypatch.setattr(autorun, "run_once", fake_run_once)
    monkeypatch.setattr(autorun, "call_optimizer_llm", fake_call)

    await autorun.main_async(argparse.Namespace(dataset="dataset.xlsx", task="code"))

    stop = json.loads((_latest_session(tmp_path, "code") / "stop.json").read_text())
    assert stop["reason"] == expected_reason
    assert len(seen_groups) == expected_count
    assert stop["no_business_learning_count"] >= cfg.no_business_learning_window


@pytest.mark.asyncio
async def test_autorun_failed_memory_uses_focused_target_not_unrelated_improvement(tmp_path, monkeypatch):
    prompt = tmp_path / "prompts" / "ocr.js"
    prompt.parent.mkdir()
    prompt.write_text("accepted", encoding="utf-8")
    focused_sample = Sample(2, "number.png", 0, "W0B053BJVLF9QJL", True)
    unrelated_dev_sample = Sample(3, "missing.png", 0, "MISSING", True)
    full_guard_sample = Sample(99, "guard.png", 0, "SAFE", True)
    samples = [focused_sample, unrelated_dev_sample, full_guard_sample]
    cfg = types.SimpleNamespace(
        dev_sample_size=2,
        ocr_concurrency=1,
        runs_dir=tmp_path / "runs",
        prompt_path=prompt,
        node_binary="node",
        ocr_runner_path="runner.js",
        max_iterations=2,
        target_business_accuracy=101.0,
        plateau_window=9,
        optimizer_provider="test",
        optimizer_model="test",
    )
    seen_groups = []

    monkeypatch.setattr(autorun.OptimizerConfig, "from_env", classmethod(lambda cls: cfg))
    monkeypatch.setattr(autorun, "load_dataset", lambda path, task: samples)
    monkeypatch.setattr(
        autorun,
        "split_samples",
        lambda items, size: DatasetSplit(dev=items[:2], full=items),
    )
    monkeypatch.setattr(
        autorun,
        "validate_prompt_file",
        lambda path, node_binary="node", task=None, baseline_source=None: None,
    )
    monkeypatch.setattr(autorun, "commit_prompt", lambda path, message: pytest.fail("should not commit"))

    async def fake_run_once(items, runner, concurrency, task):
        candidate = prompt.read_text(encoding="utf-8") != "accepted"
        results = []
        for item in items:
            actual_by_row = {
                2: "6338730878581133",
                3: "MISSING" if candidate else "",
                99: "" if candidate else "SAFE",
            }
            actual = actual_by_row[item.row_number]
            results.append(
                EvaluationResult.from_ocr_response(
                    item,
                    {
                        "status": 200,
                        "data": [{"number": actual}] if actual else [],
                        "imageStatus": ["ok"],
                    },
                )
            )
        return results

    def fake_call(provider, model, system, user):
        payload = json.loads(user)
        group_key = payload["focused_feedback_group"]["key"]
        seen_groups.append(group_key)
        return OptimizerProposal("h", "e", "r", [group_key], f"candidate-{len(seen_groups)}")

    monkeypatch.setattr(autorun, "run_once", fake_run_once)
    monkeypatch.setattr(autorun, "call_optimizer_llm", fake_call)

    await autorun.main_async(argparse.Namespace(dataset="dataset.xlsx", task="code"))

    assert seen_groups == ["wrong_code_selected_non_redeemable_number", "missing_code"]
    session = _latest_session(tmp_path, "code")
    memory = json.loads((session / "run-001/failed-strategy-memory.json").read_text())
    assert memory["entries"][0]["focused_group"] == "wrong_code_selected_non_redeemable_number"
    assert memory["entries"][0]["outcome"] == "unchanged"


@pytest.mark.asyncio
async def test_autorun_retries_once_when_optimizer_returns_invalid_prompt(tmp_path, monkeypatch):
    prompt = tmp_path / "prompts" / "ocr.js"
    prompt.parent.mkdir()
    prompt.write_text("old-valid", encoding="utf-8")
    sample = Sample(2, "a.png", 0, "Physics", True)
    cfg = types.SimpleNamespace(
        dev_sample_size=1,
        ocr_concurrency=1,
        runs_dir=tmp_path / "runs",
        prompt_path=prompt,
        node_binary="node",
        ocr_runner_path="runner.js",
        max_iterations=1,
        target_business_accuracy=101.0,
        plateau_window=3,
        optimizer_provider="test",
        optimizer_model="test",
    )
    proposals = iter(
        [
            OptimizerProposal("h1", "e1", "r1", ["row 2"], "--- prompt.js\n+++ prompt.js\n@@\n+bad"),
            OptimizerProposal("h2", "e2", "r2", ["row 2"], "new-valid"),
        ]
    )
    seen_users = []
    commits = []

    monkeypatch.setattr(autorun.OptimizerConfig, "from_env", classmethod(lambda cls: cfg))
    monkeypatch.setattr(autorun, "load_dataset", lambda path, task: [sample])
    monkeypatch.setattr(autorun, "split_samples", lambda samples, size: DatasetSplit(dev=[sample], full=[sample]))
    monkeypatch.setattr(autorun, "commit_prompt", lambda path, message: commits.append(message))

    def fake_validate(path, node_binary="node", task=None, baseline_source=None):
        if prompt.read_text(encoding="utf-8").startswith("---"):
            raise RuntimeError("prompt syntax check failed")

    async def fake_run_once(samples, runner, concurrency, task):
        actual_type = "Physics" if prompt.read_text(encoding="utf-8") == "new-valid" else "E-codes"
        return [
            EvaluationResult.from_ocr_response(
                item,
                {"status": 200, "data": [{"type": actual_type, "number": "wrong"}], "imageStatus": ["ok"]},
                task="type",
            )
            for item in samples
        ]

    def fake_call(provider, model, system, user):
        seen_users.append(user)
        return next(proposals)

    monkeypatch.setattr(autorun, "validate_prompt_file", fake_validate)
    monkeypatch.setattr(autorun, "run_once", fake_run_once)
    monkeypatch.setattr(autorun, "call_optimizer_llm", fake_call)

    await autorun.main_async(argparse.Namespace(dataset="dataset.xlsx", task="type"))

    assert len(seen_users) == 2
    retry_payload = json.loads(seen_users[1])
    assert retry_payload["task"] == "type"
    assert "PROMPT_COMPLEX" in " ".join(retry_payload["mutation_boundary"]["allowed"])
    assert retry_payload["gate_error"] == "prompt syntax check failed"
    assert prompt.read_text(encoding="utf-8") == "new-valid"
    assert commits == ["prompt(type): improve type OCR accuracy to 100.00%"]
    summary = json.loads((_latest_session(tmp_path, "type") / "run-001/summary.json").read_text())
    assert summary["task"] == "type"


@pytest.mark.asyncio
async def test_autorun_uses_type_metric_for_keep_commit_and_artifacts(tmp_path, monkeypatch):
    prompt = tmp_path / "prompts" / "ocr.js"
    prompt.parent.mkdir()
    prompt.write_text("old-valid", encoding="utf-8")
    sample = Sample(2, "a.png", 0, "Physics", True)
    cfg = types.SimpleNamespace(
        dev_sample_size=1,
        ocr_concurrency=1,
        runs_dir=tmp_path / "runs",
        prompt_path=prompt,
        node_binary="node",
        ocr_runner_path="runner.js",
        max_iterations=1,
        target_business_accuracy=101.0,
        plateau_window=3,
        optimizer_provider="test",
        optimizer_model="test",
    )
    commits = []

    monkeypatch.setattr(autorun.OptimizerConfig, "from_env", classmethod(lambda cls: cfg))
    monkeypatch.setattr(autorun, "load_dataset", lambda path, task: [sample])
    monkeypatch.setattr(autorun, "split_samples", lambda samples, size: DatasetSplit(dev=[sample], full=[sample]))
    monkeypatch.setattr(
        autorun,
        "validate_prompt_file",
        lambda path, node_binary="node", task=None, baseline_source=None: None,
    )
    monkeypatch.setattr(
        autorun,
        "call_optimizer_llm",
        lambda provider, model, system, user: OptimizerProposal("h", "e", "r", ["row 2"], "new-valid"),
    )
    monkeypatch.setattr(autorun, "commit_prompt", lambda path, message: commits.append(message))

    async def fake_run_once(samples, runner, concurrency, task):
        assert task == "type"
        actual_type = "Physics" if prompt.read_text(encoding="utf-8") == "new-valid" else "E-codes"
        return [
            EvaluationResult.from_ocr_response(
                item,
                {"status": 200, "data": [{"type": actual_type, "number": "wrong"}], "imageStatus": ["ok"]},
                task="type",
            )
            for item in samples
        ]

    monkeypatch.setattr(autorun, "run_once", fake_run_once)

    await autorun.main_async(argparse.Namespace(dataset="dataset.xlsx", task="type"))

    assert commits == ["prompt(type): improve type OCR accuracy to 100.00%"]
    run_dir = _latest_session(tmp_path, "type") / "run-001"
    summary = json.loads((run_dir / "summary.json").read_text())
    assert summary["task"] == "type"
    assert summary["type_accuracy"] == 100.0
    gate = json.loads((run_dir / "gate.json").read_text())
    assert gate["decision"] == "not_configured"
    assert gate["reason"] == "regression_not_configured"
    assert gate["checks"] == []


@pytest.mark.asyncio
async def test_autorun_records_gate_failure_without_running_dev_or_full(tmp_path, monkeypatch):
    prompt = tmp_path / "prompts" / "ocr.js"
    prompt.parent.mkdir()
    prompt.write_text("old-valid", encoding="utf-8")
    sample = Sample(2, "a.png", 0, "A", True)
    cfg = types.SimpleNamespace(
        dev_sample_size=1,
        ocr_concurrency=1,
        runs_dir=tmp_path / "runs",
        prompt_path=prompt,
        node_binary="node",
        ocr_runner_path="runner.js",
        max_iterations=1,
        target_business_accuracy=101.0,
        plateau_window=3,
        optimizer_provider="test",
        optimizer_model="test",
    )
    run_prompts = []
    proposals = iter(
        [
            OptimizerProposal("h1", "e1", "r1", ["row 2"], "bad-boundary"),
            OptimizerProposal("h2", "e2", "r2", ["row 2"], "bad-boundary"),
        ]
    )

    monkeypatch.setattr(autorun.OptimizerConfig, "from_env", classmethod(lambda cls: cfg))
    monkeypatch.setattr(autorun, "load_dataset", lambda path, task: [sample])
    monkeypatch.setattr(autorun, "split_samples", lambda samples, size: DatasetSplit(dev=[sample], full=[sample]))
    monkeypatch.setattr(autorun, "call_optimizer_llm", lambda provider, model, system, user: next(proposals))
    monkeypatch.setattr(autorun, "commit_prompt", lambda path, message: pytest.fail("should not commit"))

    def fake_validate(path, node_binary="node", task=None, baseline_source=None):
        if prompt.read_text(encoding="utf-8") == "bad-boundary":
            raise RuntimeError("code task cannot change protected exports")

    async def fake_run_once(samples, runner, concurrency, task):
        run_prompts.append(prompt.read_text(encoding="utf-8"))
        return [
            EvaluationResult.from_ocr_response(
                item,
                {"status": 200, "data": [{"number": "A"}], "imageStatus": ["ok"]},
            )
            for item in samples
        ]

    monkeypatch.setattr(autorun, "validate_prompt_file", fake_validate)
    monkeypatch.setattr(autorun, "run_once", fake_run_once)

    await autorun.main_async(argparse.Namespace(dataset="dataset.xlsx", task="code"))

    assert run_prompts == ["old-valid"]
    run_dir = _latest_session(tmp_path, "code") / "run-001"
    summary = json.loads((run_dir / "summary.json").read_text())
    response = json.loads((run_dir / "optimizer-response.json").read_text())
    assert summary["phase"] == "gate_failed"
    assert summary["task"] == "code"
    assert "code task cannot change protected exports" in response["gate_error"]


@pytest.mark.asyncio
async def test_autorun_loads_regression_dataset_with_selected_task_schema(tmp_path, monkeypatch):
    prompt = tmp_path / "prompts" / "ocr.js"
    prompt.parent.mkdir()
    prompt.write_text("valid", encoding="utf-8")
    sample = Sample(2, "a.png", 0, "Physics", True)
    cfg = types.SimpleNamespace(
        dev_sample_size=1,
        ocr_concurrency=1,
        runs_dir=tmp_path / "runs",
        prompt_path=prompt,
        node_binary="node",
        ocr_runner_path="runner.js",
        max_iterations=0,
        target_business_accuracy=101.0,
        plateau_window=3,
        optimizer_provider="test",
        optimizer_model="test",
    )
    loaded = []

    monkeypatch.setattr(autorun.OptimizerConfig, "from_env", classmethod(lambda cls: cfg))

    def fake_load_dataset(path, task):
        loaded.append((path, task))
        return [sample]

    async def fake_run_once(samples, runner, concurrency, task):
        return [
            EvaluationResult.from_ocr_response(
                item,
                {"status": 200, "data": [{"type": "Physics", "number": "wrong"}], "imageStatus": ["ok"]},
                task="type",
            )
            for item in samples
        ]

    monkeypatch.setattr(autorun, "load_dataset", fake_load_dataset)
    monkeypatch.setattr(autorun, "split_samples", lambda samples, size: DatasetSplit(dev=[sample], full=[sample]))
    monkeypatch.setattr(
        autorun,
        "validate_prompt_file",
        lambda path, node_binary="node", task=None, baseline_source=None: None,
    )
    monkeypatch.setattr(autorun, "run_once", fake_run_once)

    await autorun.main_async(
        argparse.Namespace(dataset="dataset.xlsx", regression_dataset="regression.xlsx", task="type")
    )

    assert loaded == [("dataset.xlsx", "type"), ("regression.xlsx", "type")]


@pytest.mark.asyncio
async def test_autorun_fails_fast_when_regression_dataset_is_malformed(tmp_path, monkeypatch):
    prompt = tmp_path / "prompts" / "ocr.js"
    prompt.parent.mkdir()
    prompt.write_text("valid", encoding="utf-8")
    sample = Sample(2, "a.png", 0, "A", True)
    cfg = types.SimpleNamespace(
        dev_sample_size=1,
        ocr_concurrency=1,
        runs_dir=tmp_path / "runs",
        prompt_path=prompt,
        node_binary="node",
        ocr_runner_path="runner.js",
        max_iterations=0,
        target_business_accuracy=101.0,
        plateau_window=3,
        optimizer_provider="test",
        optimizer_model="test",
    )

    monkeypatch.setattr(autorun.OptimizerConfig, "from_env", classmethod(lambda cls: cfg))

    def fake_load_dataset(path, task):
        if path == "regression.xlsx":
            raise ValueError("missing required columns: md5_card_number")
        return [sample]

    monkeypatch.setattr(autorun, "load_dataset", fake_load_dataset)
    monkeypatch.setattr(autorun, "run_once", lambda *args, **kwargs: pytest.fail("should fail before OCR"))

    with pytest.raises(ValueError, match="missing required columns: md5_card_number"):
        await autorun.main_async(
            argparse.Namespace(dataset="dataset.xlsx", regression_dataset="regression.xlsx", task="code")
        )


@pytest.mark.asyncio
async def test_autorun_discards_full_improvement_when_regression_gate_fails(tmp_path, monkeypatch):
    prompt = tmp_path / "prompts" / "ocr.js"
    prompt.parent.mkdir()
    prompt.write_text("accepted", encoding="utf-8")
    main_sample = Sample(2, "main.png", 0, "MAIN", True)
    regression_sample = Sample(99, "guard.png", 0, "SAFE", True)
    cfg = types.SimpleNamespace(
        dev_sample_size=1,
        ocr_concurrency=1,
        runs_dir=tmp_path / "runs",
        prompt_path=prompt,
        node_binary="node",
        ocr_runner_path="runner.js",
        max_iterations=2,
        target_business_accuracy=101.0,
        plateau_window=3,
        optimizer_provider="test",
        optimizer_model="test",
    )
    proposals = iter(
        [
            OptimizerProposal("h1", "e1", "r1", ["row 99"], "candidate"),
            OptimizerProposal("h2", "e2", "r2", ["row 99"], "accepted"),
        ]
    )
    seen_accuracies = []
    run_labels = []
    commits = []

    monkeypatch.setattr(autorun.OptimizerConfig, "from_env", classmethod(lambda cls: cfg))

    def fake_load_dataset(path, task):
        return [regression_sample] if path == "regression.xlsx" else [main_sample]

    async def fake_run_once(samples, runner, concurrency, task):
        sample = samples[0]
        prompt_text = prompt.read_text(encoding="utf-8")
        is_regression = sample.row_number == regression_sample.row_number
        run_labels.append(("regression" if is_regression else "main", prompt_text))
        if is_regression:
            actual = "SAFE" if prompt_text == "accepted" else "MISS"
        else:
            actual = "MAIN" if prompt_text == "candidate" else "MISS"
        return [
            EvaluationResult.from_ocr_response(
                sample,
                {"status": 200, "data": [{"number": actual}], "imageStatus": ["ok"]},
            )
        ]

    def fake_call(provider, model, system, user):
        seen_accuracies.append(json.loads(user)["summary"]["business_accuracy"])
        return next(proposals)

    monkeypatch.setattr(autorun, "load_dataset", fake_load_dataset)
    monkeypatch.setattr(autorun, "split_samples", lambda samples, size: DatasetSplit(dev=samples, full=samples))
    monkeypatch.setattr(
        autorun,
        "validate_prompt_file",
        lambda path, node_binary="node", task=None, baseline_source=None: None,
    )
    monkeypatch.setattr(autorun, "run_once", fake_run_once)
    monkeypatch.setattr(autorun, "call_optimizer_llm", fake_call)
    monkeypatch.setattr(autorun, "commit_prompt", lambda path, message: commits.append(message))

    await autorun.main_async(
        argparse.Namespace(dataset="dataset.xlsx", regression_dataset="regression.xlsx", task="code")
    )

    assert commits == []
    assert prompt.read_text(encoding="utf-8") == "accepted"
    assert seen_accuracies == [0.0, 0.0]
    assert run_labels == [
        ("main", "accepted"),
        ("regression", "accepted"),
        ("main", "candidate"),
        ("main", "candidate"),
        ("regression", "candidate"),
    ]
    run_dir = _latest_session(tmp_path, "code") / "run-001"
    gate = json.loads((run_dir / "gate.json").read_text())
    full_summary = json.loads((run_dir / "summary.json").read_text())
    regression_summary = json.loads((run_dir / "regression-summary.json").read_text())
    response = json.loads((run_dir / "optimizer-response.json").read_text())

    assert gate["decision"] == "discard"
    assert (run_dir / "prompt-before.js").read_text() == "accepted"
    assert (run_dir / "prompt-after.js").read_text() == "candidate"
    assert "candidate" in (run_dir / "prompt.diff").read_text()
    assert full_summary["phase"] == "full"
    assert full_summary["business_accuracy"] == 100.0
    assert _xlsx_headers(run_dir / "results.xlsx") == CODE_RESULT_HEADERS
    assert regression_summary["phase"] == "regression"
    assert regression_summary["business_accuracy"] == 0.0
    assert _xlsx_headers(run_dir / "regression-results.xlsx") == CODE_RESULT_HEADERS
    assert response["target_failures"] == ["row 99"]


@pytest.mark.asyncio
async def test_autorun_writes_type_regression_failure_artifacts(tmp_path, monkeypatch):
    prompt = tmp_path / "prompts" / "ocr.js"
    prompt.parent.mkdir()
    prompt.write_text("accepted", encoding="utf-8")
    main_sample = Sample(2, "main.png", 0, "Physics", True)
    regression_sample = Sample(99, "guard.png", 0, "Physics", True)
    cfg = types.SimpleNamespace(
        dev_sample_size=1,
        ocr_concurrency=1,
        runs_dir=tmp_path / "runs",
        prompt_path=prompt,
        node_binary="node",
        ocr_runner_path="runner.js",
        max_iterations=1,
        target_business_accuracy=101.0,
        plateau_window=3,
        optimizer_provider="test",
        optimizer_model="test",
    )

    monkeypatch.setattr(autorun.OptimizerConfig, "from_env", classmethod(lambda cls: cfg))
    monkeypatch.setattr(
        autorun,
        "call_optimizer_llm",
        lambda provider, model, system, user: OptimizerProposal(
            "h",
            "e",
            "r",
            ["row 99"],
            "candidate",
        ),
    )
    monkeypatch.setattr(autorun, "commit_prompt", lambda path, message: pytest.fail("should not commit"))

    def fake_load_dataset(path, task):
        return [regression_sample] if path == "regression.xlsx" else [main_sample]

    async def fake_run_once(samples, runner, concurrency, task):
        sample = samples[0]
        prompt_text = prompt.read_text(encoding="utf-8")
        if sample.row_number == regression_sample.row_number:
            actual_type = "Physics" if prompt_text == "accepted" else "E-codes"
        else:
            actual_type = "Physics" if prompt_text == "candidate" else "E-codes"
        return [
            EvaluationResult.from_ocr_response(
                sample,
                {"status": 200, "data": [{"type": actual_type, "number": "wrong"}], "imageStatus": ["ok"]},
                task="type",
            )
        ]

    monkeypatch.setattr(autorun, "load_dataset", fake_load_dataset)
    monkeypatch.setattr(autorun, "split_samples", lambda samples, size: DatasetSplit(dev=samples, full=samples))
    monkeypatch.setattr(
        autorun,
        "validate_prompt_file",
        lambda path, node_binary="node", task=None, baseline_source=None: None,
    )
    monkeypatch.setattr(autorun, "run_once", fake_run_once)

    await autorun.main_async(
        argparse.Namespace(dataset="dataset.xlsx", regression_dataset="regression.xlsx", task="type")
    )

    run_dir = _latest_session(tmp_path, "type") / "run-001"
    gate = json.loads((run_dir / "gate.json").read_text())
    regression_summary = json.loads((run_dir / "regression-summary.json").read_text())

    assert prompt.read_text(encoding="utf-8") == "accepted"
    assert gate["decision"] == "discard"
    assert regression_summary["phase"] == "regression"
    assert regression_summary["type_accuracy"] == 0.0
    assert _xlsx_headers(run_dir / "regression-results.xlsx") == TYPE_RESULT_HEADERS


@pytest.mark.asyncio
async def test_autorun_updates_regression_baseline_after_accepted_candidate(tmp_path, monkeypatch):
    prompt = tmp_path / "prompts" / "ocr.js"
    prompt.parent.mkdir()
    prompt.write_text("accepted", encoding="utf-8")
    main_sample = Sample(2, "main.png", 0, "A\nB", True)
    regression_sample = Sample(99, "guard.png", 0, "SAFE\nGUARD", True)
    cfg = types.SimpleNamespace(
        dev_sample_size=1,
        ocr_concurrency=1,
        runs_dir=tmp_path / "runs",
        prompt_path=prompt,
        node_binary="node",
        ocr_runner_path="runner.js",
        max_iterations=2,
        target_business_accuracy=101.0,
        plateau_window=3,
        optimizer_provider="test",
        optimizer_model="test",
    )
    proposals = iter(
        [
            OptimizerProposal("h1", "e1", "r1", ["row 2"], "candidate-one"),
            OptimizerProposal("h2", "e2", "r2", ["row 2"], "candidate-two"),
        ]
    )
    seen_accuracies = []
    seen_payloads = []
    commits = []

    monkeypatch.setattr(autorun.OptimizerConfig, "from_env", classmethod(lambda cls: cfg))

    def fake_load_dataset(path, task):
        return [regression_sample] if path == "regression.xlsx" else [main_sample]

    async def fake_run_once(samples, runner, concurrency, task):
        sample = samples[0]
        prompt_text = prompt.read_text(encoding="utf-8")
        if sample.row_number == regression_sample.row_number:
            actual = "SAFE" if prompt_text == "candidate-one" else ""
        elif prompt_text == "candidate-one":
            actual = "A"
        elif prompt_text == "candidate-two":
            actual = "A\nB"
        else:
            actual = ""
        return [
            EvaluationResult.from_ocr_response(
                sample,
                {"status": 200, "data": [{"number": actual}], "imageStatus": ["ok"]},
            )
        ]

    def fake_call(provider, model, system, user):
        payload = json.loads(user)
        seen_payloads.append(payload)
        seen_accuracies.append(payload["summary"]["business_accuracy"])
        return next(proposals)

    monkeypatch.setattr(autorun, "load_dataset", fake_load_dataset)
    monkeypatch.setattr(autorun, "split_samples", lambda samples, size: DatasetSplit(dev=samples, full=samples))
    monkeypatch.setattr(
        autorun,
        "validate_prompt_file",
        lambda path, node_binary="node", task=None, baseline_source=None: None,
    )
    monkeypatch.setattr(autorun, "run_once", fake_run_once)
    monkeypatch.setattr(autorun, "call_optimizer_llm", fake_call)
    monkeypatch.setattr(autorun, "commit_prompt", lambda path, message: commits.append(message))

    await autorun.main_async(
        argparse.Namespace(dataset="dataset.xlsx", regression_dataset="regression.xlsx", task="code")
    )

    assert commits == ["prompt(code): improve code OCR accuracy to 50.00%"]
    assert prompt.read_text(encoding="utf-8") == "candidate-one"
    assert seen_accuracies == [0.0, 50.0]
    assert "candidate_evaluation_delta_summary" not in seen_payloads[1]
    session = _latest_session(tmp_path, "code")
    first_delta = json.loads((session / "run-001/dev-delta.json").read_text())
    second_delta = json.loads((session / "run-002/dev-delta.json").read_text())
    assert first_delta["improved_business_rows"][0]["accepted_business_correct"] == 0
    assert first_delta["improved_business_rows"][0]["candidate_business_correct"] == 1
    assert second_delta["improved_business_rows"][0]["accepted_business_correct"] == 1
    assert second_delta["improved_business_rows"][0]["candidate_business_correct"] == 2
